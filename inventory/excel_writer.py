# excel_writer.py (updated)
import pandas as pd
from utils import logger
from openpyxl.styles import PatternFill

def write_to_excel(dataframes_dict, output_file="aws_inventory.xlsx"):
    try:
        logger.info(f"Writing inventory to {output_file}")

        with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
            for sheet_name, df in dataframes_dict.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)
                workbook = writer.book
                worksheet = writer.sheets[sheet_name]

                # Highlight missing tags in red
                if "Tags" in df.columns:
                    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    for row_idx, cell_value in enumerate(df["Tags"], start=2):  # +2 because header=1
                        if not cell_value:
                            worksheet.cell(row=row_idx, column=df.columns.get_loc("Tags")+1).fill = red_fill

        logger.info("Excel report created successfully.")

    except Exception as e:
        logger.error(f"Excel writing error: {e}")

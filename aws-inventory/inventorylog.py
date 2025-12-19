import boto3
import logging
from botocore.exceptions import ClientError
from openpyxl import Workbook

# -----------------------------------
# LOGGING CONFIGURATION
# -----------------------------------
logging.basicConfig(
    filename="aws_inventory.log",
    level=logging.INFO,
    format="%(asctime)s - %(levelname)s - %(message)s"
)

console = logging.StreamHandler()
console.setLevel(logging.INFO)
formatter = logging.Formatter("%(asctime)s - %(levelname)s - %(message)s")
console.setFormatter(formatter)
logging.getLogger("").addHandler(console)

logging.info("AWS Inventory Script Started")

# -----------------------------------
# GET ALL REGIONS
# -----------------------------------
def get_all_regions():
    try:
        logging.info("Fetching all AWS regions")
        ec2 = boto3.client("ec2")
        response = ec2.describe_regions(AllRegions=True)
        regions = [r["RegionName"] for r in response["Regions"]]
        logging.info(f"Regions fetched: {regions}")
        return regions
    except ClientError as e:
        logging.error(f"Error fetching regions: {e}")
        return []
    except Exception as e:
        logging.error(f"Unexpected error fetching regions: {e}")
        return []


# -----------------------------------
# EC2 INSTANCES
# -----------------------------------
def get_ec2_instances(region):
    logging.info(f"Fetching EC2 instances in region: {region}")
    try:
        ec2 = boto3.client("ec2", region_name=region)

        try:
            response = ec2.describe_instances()
        except ClientError as e:
            code = e.response.get("Error", {}).get("Code")
            if code in ["AuthFailure", "UnrecognizedClientException", "UnauthorizedOperation"]:
                logging.warning(f"EC2 access denied for region {region}: {code}")
                return []
            else:
                logging.error(f"EC2 error in {region}: {e}")
                return []
        except Exception as e:
            logging.error(f"Unexpected EC2 error in {region}: {e}")
            return []

        instances = []
        for reservation in response.get("Reservations", []):
            for instance in reservation.get("Instances", []):
                instances.append({
                    "Region": region,
                    "InstanceId": instance.get("InstanceId"),
                    "InstanceType": instance.get("InstanceType"),
                    "State": instance.get("State", {}).get("Name"),
                    "PrivateIP": instance.get("PrivateIpAddress"),
                    "PublicIP": instance.get("PublicIpAddress"),
                    "AZ": instance.get("Placement", {}).get("AvailabilityZone"),
                })

        logging.info(f"EC2 instances found in {region}: {len(instances)}")
        return instances

    except Exception as e:
        logging.error(f"Failed to fetch EC2 instances in {region}: {e}")
        return []


# -----------------------------------
# S3 BUCKETS
# -----------------------------------
def get_s3_buckets():
    logging.info("Fetching S3 buckets")
    try:
        s3 = boto3.client("s3")

        try:
            response = s3.list_buckets()
        except ClientError as e:
            logging.error(f"S3 client error: {e}")
            return []
        except Exception as e:
            logging.error(f"Unexpected S3 error: {e}")
            return []

        buckets = []
        for bucket in response.get("Buckets", []):
            buckets.append({
                "BucketName": bucket.get("Name"),
                "CreationDate": str(bucket.get("CreationDate")),
            })

        logging.info(f"S3 buckets found: {len(buckets)}")
        return buckets

    except Exception as e:
        logging.error(f"Failed to fetch S3 buckets: {e}")
        return []


# -----------------------------------
# LAMBDA FUNCTIONS
# -----------------------------------
def get_lambda_functions(region):
    logging.info(f"Fetching Lambda functions in region: {region}")
    try:
        lam = boto3.client("lambda", region_name=region)

        try:
            response = lam.list_functions()
        except ClientError as e:
            code = e.response.get("Error", {}).get("Code")
            if code in ["AuthFailure", "UnrecognizedClientException", "UnauthorizedOperation"]:
                logging.warning(f"Lambda access denied in {region}: {code}")
                return []
            else:
                logging.error(f"Lambda error in {region}: {e}")
                return []
        except Exception as e:
            logging.error(f"Unexpected Lambda error in {region}: {e}")
            return []

        functions = []
        for fn in response.get("Functions", []):
            functions.append({
                "Region": region,
                "FunctionName": fn.get("FunctionName"),
                "Runtime": fn.get("Runtime"),
                "Memory": fn.get("MemorySize"),
                "Timeout": fn.get("Timeout"),
            })

        logging.info(f"Lambda functions found in {region}: {len(functions)}")
        return functions

    except Exception as e:
        logging.error(f"Failed to fetch Lambda functions in {region}: {e}")
        return []


# -----------------------------------
# WRITE TO EXCEL
# -----------------------------------
def write_to_excel(ec2_data, s3_data, lambda_data):
    logging.info("Writing data to Excel file aws_inventory2.xlsx")
    try:
        wb = Workbook()

        # EC2 Sheet
        try:
            ws1 = wb.active
            ws1.title = "EC2"
            ws1.append(["Region", "InstanceId", "InstanceType", "State", "PrivateIP", "PublicIP", "AZ"])
            for row in ec2_data:
                ws1.append(list(row.values()))
        except Exception as e:
            logging.error(f"Error writing EC2 sheet: {e}")

        # S3 Sheet
        try:
            ws2 = wb.create_sheet("S3")
            ws2.append(["BucketName", "CreationDate"])
            for row in s3_data:
                ws2.append(list(row.values()))
        except Exception as e:
            logging.error(f"Error writing S3 sheet: {e}")

        # Lambda sheet
        try:
            ws3 = wb.create_sheet("Lambda")
            ws3.append(["Region", "FunctionName", "Runtime", "Memory", "Timeout"])
            for row in lambda_data:
                ws3.append(list(row.values()))
        except Exception as e:
            logging.error(f"Error writing Lambda sheet: {e}")

        wb.save("aws_inventory1.xlsx")
        logging.info("Excel export completed successfully")

    except Exception as e:
        logging.error(f"Failed to save Excel file: {e}")


# -----------------------------------
# MAIN
# -----------------------------------
def main():
    logging.info("Script execution started")

    try:
        regions = get_all_regions()
    except Exception as e:
        logging.error(f"Cannot get regions, exiting: {e}")
        return

    all_ec2 = []
    all_lambda = []

    for region in regions:
        try:
            all_ec2.extend(get_ec2_instances(region))
        except Exception as e:
            logging.error(f"Error processing EC2 in {region}: {e}")

        try:
            all_lambda.extend(get_lambda_functions(region))
        except Exception as e:
            logging.error(f"Error processing Lambda in {region}: {e}")

    try:
        s3 = get_s3_buckets()
    except Exception as e:
        logging.error(f"Error processing S3: {e}")
        s3 = []

    try:
        write_to_excel(all_ec2, s3, all_lambda)
    except Exception as e:
        logging.error(f"Error writing Excel: {e}")

    logging.info("Script execution finished")


if __name__ == "__main__":
    main()

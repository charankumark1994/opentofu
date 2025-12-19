import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook

def get_all_regions():
    try:
        ec2 = boto3.client("ec2")
        response = ec2.describe_regions(AllRegions=True)
        return [r["RegionName"] for r in response["Regions"]]
    except ClientError as e:
        print("Error fetching regions:", e)
        return []

def get_ec2_instances(region):
    try:
        ec2 = boto3.client("ec2", region_name=region)
        response = ec2.describe_instances()

        instances = []
        for reservation in response.get("Reservations", []):
            for instance in reservation.get("Instances", []):
                instances.append({
                    "Region": region,
                    "InstanceId": instance.get("InstanceId"),
                    "InstanceType": instance.get("InstanceType"),
                    "State": instance.get("State", {}).get("Name"),
                    "PrivateIP": instance.get("PrivateIpAddress"),
                    "PublicIP": instance.get("PublicIpAddress"),
                    "AZ": instance.get("Placement", {}).get("AvailabilityZone"),
                })
        return instances
    except ClientError as e:
        print(f"Error fetching EC2 in {region}:", e)
        return []

def get_s3_buckets():
    try:
        s3 = boto3.client("s3")
        response = s3.list_buckets()

        buckets = []
        for bucket in response.get("Buckets", []):
            buckets.append({
                "BucketName": bucket.get("Name"),
                "CreationDate": str(bucket.get("CreationDate")),
            })
        return buckets
    except ClientError as e:
        print("Error fetching S3 buckets:", e)
        return []

def get_lambda_functions(region):
    try:
        lam = boto3.client("lambda", region_name=region)
        functions = []

        paginator = lam.get_paginator("list_functions")
        for page in paginator.paginate():
            for fn in page.get("Functions", []):
                functions.append({
                    "Region": region,
                    "FunctionName": fn.get("FunctionName"),
                    "Runtime": fn.get("Runtime"),
                    "Memory": fn.get("MemorySize"),
                    "Timeout": fn.get("Timeout"),
                })
        return functions
    except ClientError as e:
        print(f"Error fetching Lambda in {region}:", e)
        return []

def write_to_excel(ec2_data, s3_data, lambda_data):
    try:
        wb = Workbook()

        ws1 = wb.active
        ws1.title = "EC2"
        ws1.append(["Region", "InstanceId", "InstanceType", "State", "PrivateIP", "PublicIP", "AZ"])
        for row in ec2_data:
            ws1.append(list(row.values()))

        ws2 = wb.create_sheet("S3")
        ws2.append(["BucketName", "CreationDate"])
        for row in s3_data:
            ws2.append(list(row.values()))

        ws3 = wb.create_sheet("Lambda")
        ws3.append(["Region", "FunctionName", "Runtime", "Memory", "Timeout"])
        for row in lambda_data:
            ws3.append(list(row.values()))

        wb.save("aws_inventory.xlsx")
        print("Export completed: aws_inventory.xlsx")
    except Exception as e:
        print("Error writing to Excel:", e)

def main():
    regions = get_all_regions()

    all_ec2 = []
    all_lambda = []

    for region in regions:
        all_ec2.extend(get_ec2_instances(region))
        all_lambda.extend(get_lambda_functions(region))

    s3 = get_s3_buckets()

    write_to_excel(all_ec2, s3, all_lambda)

if __name__ == "__main__":
    main()

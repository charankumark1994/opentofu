import boto3
from botocore.exceptions import ClientError
from openpyxl import Workbook

def get_all_regions():
    try:
        ec2 = boto3.client("ec2")
        response = ec2.describe_regions(AllRegions=True)
        return [r["RegionName"] for r in response["Regions"]]
    except ClientError as e:
        print("Error fetching regions:", e)
        return []
    except Exception as e:
        print("Unexpected error fetching regions:", e)
        return []

def get_ec2_instances(region):
    try:
        ec2 = boto3.client("ec2", region_name=region)
        try:
            response = ec2.describe_instances()
        except ClientError as e:
            code = e.response.get("Error", {}).get("Code")
            if code in ["AuthFailure", "UnrecognizedClientException", "UnauthorizedOperation"]:
                return []
            else:
                print(f"EC2 error in {region}:", e)
                return []
        except Exception as e:
            print(f"Unexpected EC2 error in {region}:", e)
            return []

        instances = []
        for reservation in response.get("Reservations", []):
            for instance in reservation.get("Instances", []):
                instances.append({
                    "Region": region,
                    "InstanceId": instance.get("InstanceId"),
                    "InstanceType": instance.get("InstanceType"),
                    "State": instance.get("State", {}).get("Name"),
                    "PrivateIP": instance.get("PrivateIpAddress"),
                    "PublicIP": instance.get("PublicIpAddress"),
                    "AZ": instance.get("Placement", {}).get("AvailabilityZone"),
                })
        return instances
    except Exception as e:
        print(f"Failed to fetch EC2 instances in {region}:", e)
        return []

def get_s3_buckets():
    try:
        s3 = boto3.client("s3")
        try:
            response = s3.list_buckets()
        except ClientError as e:
            print("S3 error:", e)
            return []
        except Exception as e:
            print("Unexpected S3 error:", e)
            return []

        buckets = []
        for bucket in response.get("Buckets", []):
            buckets.append({
                "BucketName": bucket.get("Name"),
                "CreationDate": str(bucket.get("CreationDate")),
            })
        return buckets
    except Exception as e:
        print("Failed to fetch S3 buckets:", e)
        return []

def get_lambda_functions(region):
    try:
        lam = boto3.client("lambda", region_name=region)
        try:
            response = lam.list_functions()
        except ClientError as e:
            code = e.response.get("Error", {}).get("Code")
            if code in ["AuthFailure", "UnrecognizedClientException", "UnauthorizedOperation"]:
                return []
            else:
                print(f"Lambda error in {region}:", e)
                return []
        except Exception as e:
            print(f"Unexpected Lambda error in {region}:", e)
            return []

        functions = []
        for fn in response.get("Functions", []):
            functions.append({
                "Region": region,
                "FunctionName": fn.get("FunctionName"),
                "Runtime": fn.get("Runtime"),
                "Memory": fn.get("MemorySize"),
                "Timeout": fn.get("Timeout"),
            })
        return functions
    except Exception as e:
        print(f"Failed to fetch Lambda functions in {region}:", e)
        return []

def write_to_excel(ec2_data, s3_data, lambda_data):
    try:
        wb = Workbook()

        # EC2 Sheet
        try:
            ws1 = wb.active
            ws1.title = "EC2"
            ws1.append(["Region", "InstanceId", "InstanceType", "State", "PrivateIP", "PublicIP", "AZ"])
            for row in ec2_data:
                ws1.append(list(row.values()))
        except Exception as e:
            print("Error writing EC2 data to Excel:", e)

        # S3 Sheet
        try:
            ws2 = wb.create_sheet("S3")
            ws2.append(["BucketName", "CreationDate"])
            for row in s3_data:
                ws2.append(list(row.values()))
        except Exception as e:
            print("Error writing S3 data to Excel:", e)

        # Lambda Sheet
        try:
            ws3 = wb.create_sheet("Lambda")
            ws3.append(["Region", "FunctionName", "Runtime", "Memory", "Timeout"])
            for row in lambda_data:
                ws3.append(list(row.values()))
        except Exception as e:
            print("Error writing Lambda data to Excel:", e)

        wb.save("aws_inventory1.xlsx")
        print("Export completed: aws_inventory1.xlsx")
    except Exception as e:
        print("Failed to save Excel file:", e)

def main():
    try:
        regions = get_all_regions()
    except Exception as e:
        print("Cannot get regions, exiting:", e)
        return

    all_ec2 = []
    all_lambda = []

    for region in regions:
        try:
            ec2_instances = get_ec2_instances(region)
            all_ec2.extend(ec2_instances)
        except Exception as e:
            print(f"Error processing EC2 in {region}:", e)

        try:
            lambda_functions = get_lambda_functions(region)
            all_lambda.extend(lambda_functions)
        except Exception as e:
            print(f"Error processing Lambda in {region}:", e)

    try:
        s3 = get_s3_buckets()
    except Exception as e:
        print("Error processing S3:", e)
        s3 = []

    try:
        write_to_excel(all_ec2, s3, all_lambda)
    except Exception as e:
        print("Error writing Excel:", e)

if __name__ == "__main__":
    main()
